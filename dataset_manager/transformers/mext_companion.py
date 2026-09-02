"""MEXT's companion composition books: amino acids, fatty acids, carbohydrates.

The main table gives one number for protein, one for fat, one for
carbohydrate. MEXT publishes three further books that break each of those
open — the eighteen amino acids, the individual fatty acids with their
saturated/monounsaturated/polyunsaturated and n-3/n-6 totals, and the sugars,
starch and organic acids behind "carbohydrate". Same licence, same 食品番号,
and all of it was sitting unused.

Every file shares one layout, so one reader serves all ten:

    row n     Japanese component names
    row n+1   成分識別子 (the codes)
    row n+2   単位
    row n+3.. data, one row per food

Values carry the same parenthesis and Tr conventions as the main table, so
clean_value() decides what each figure is worth. Codes the main table already
owns (water, protein, fat) repeat here as context columns and are skipped.
"""
from pathlib import Path

import openpyxl

from .mext import NUTRIENT_COLS, clean_value

CODE_ROW_LABEL = "成分識別子"
UNIT_ROW_LABEL = "単位"


def _label(cell):
    """Header labels are spaced for print — 「単　位」 — so compare without it."""
    return "".join(str(cell).split()).replace("　", "") if cell is not None else ""
GROUP_COL, NUMBER_COL, INDEX_COL, NAME_COL = 0, 1, 2, 3

# Codes the main table is the authority for. They appear in these books only
# to orient the reader, and storing them twice would put two rows for the same
# nutrient on a food page.
MAIN_TABLE_CODES = {code for _c, code, _n, _u in NUTRIENT_COLS}


def _header_rows(rows):
    """(names, codes, units, first_data_index) or None if the layout is not
    what we expect — better to skip a file than to import a guess."""
    for i, row in enumerate(rows[:20]):
        if any(_label(c) == CODE_ROW_LABEL for c in row):
            names = rows[i - 1] if i else ()
            units = rows[i + 1] if len(rows) > i + 1 else ()
            if not any(_label(c) == UNIT_ROW_LABEL for c in units):
                units = ()
            return names, row, units, i + (2 if units else 1)
    return None


def read_file(path):
    """Yield (食品番号, code, japanese_name, unit, amount, quality)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["表全体"] if "表全体" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = _header_rows(rows)
    if not header:
        raise ValueError(f"{path}: no '{CODE_ROW_LABEL}' row — layout changed?")
    names, codes, units, start = header

    columns = []
    for i, raw in enumerate(codes):
        code = str(raw).strip() if raw is not None else ""
        if not code or _label(raw) in (CODE_ROW_LABEL, UNIT_ROW_LABEL) or code in MAIN_TABLE_CODES:
            continue
        jp = str(names[i]).replace("\n", "").strip() if i < len(names) and names[i] else code
        unit = str(units[i]).strip() if units and i < len(units) and units[i] else ""
        columns.append((i, code, jp, unit))

    for row in rows[start:]:
        if not row or row[NUMBER_COL] is None:
            continue
        food_code = str(row[NUMBER_COL]).strip()
        if not food_code or not food_code[0].isdigit():
            continue
        for i, code, jp, unit in columns:
            if i >= len(row):
                continue
            amount, quality = clean_value(row[i])
            if amount is None:
                continue
            yield food_code, code, jp, unit, amount, quality


def ingest(conn, path, source="MEXT Standard Tables"):
    """Load one book into `nutrients`, replacing anything a previous run of
    the same file left behind so re-running changes nothing."""
    by_code = {
        r[0]: r[1] for r in conn.execute(
            "SELECT source_url, id FROM items WHERE source = ? AND source_url LIKE 'mext_%'",
            (source,))
    }
    rows, codes_seen, missing = [], set(), set()
    for food_code, code, jp, unit, amount, quality in read_file(path):
        item_id = by_code.get(f"mext_{food_code}")
        if item_id is None:
            missing.add(food_code)
            continue
        codes_seen.add(code)
        rows.append((item_id, code, jp, unit, amount, quality))

    if codes_seen:
        marks = ",".join("?" * len(codes_seen))
        conn.execute(f"DELETE FROM nutrients WHERE code IN ({marks})", tuple(codes_seen))
    conn.executemany(
        "INSERT INTO nutrients (item_id, code, name, unit, amount, quality)"
        " VALUES (?, ?, ?, ?, ?, ?)", rows)
    conn.commit()
    return {"file": Path(path).name, "values": len(rows), "codes": len(codes_seen),
            "unmatched_foods": len(missing)}
