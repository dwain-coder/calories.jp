"""Apply MEXT's 正誤表 to the stored composition values.

    uv run python tools/apply_mext_errata.py --dry-run
    uv run python tools/apply_mext_errata.py

MEXT publishes corrections to the composition tables after release, and this
site had never applied them: some stored figures are ones the ministry has
since withdrawn. The errata corrects two different things, and both matter
here — the value itself (まあじ フライ 卵液 7.5 -> 9.2) and what the value is
worth (食物繊維総量 3.6 -> (3.6), i.e. the same number, reclassified as an
estimate; レチノール活性当量 0 -> Tr, a zero that is really a trace).

Keyed on 食品番号, which items.source_url now carries as "mext_01001".

Two sheets are read:
  本表第2章 — one row per corrected value
  本表      — whole replacement rows, in main-table layout
Remarks (備考) corrections are skipped: that column is not stored.
"""
import argparse
import os
import sqlite3
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from dataset_manager.transformers.mext import NUTRIENT_COLS, clean_value  # noqa: E402

DB_PATH = os.environ.get("DATABASE_PATH", "data/metadata/dataset_manager.db")
XLSX = Path("data/raw/mext_errata_2023/errata.xlsx")
SKIP_ITEMS = ("備考",)          # remarks text, not a stored value

# 項目等 in the errata is a Japanese nutrient name; the transformer's spec has
# the same names against their codes.
NAME_TO_CODE = {}
for _col, _code, _jp, _unit in NUTRIENT_COLS:
    NAME_TO_CODE.setdefault(_jp, _code)
# Names the errata writes slightly differently from the main table header.
def _norm(label):
    """The errata breaks long labels across lines and spells the units out."""
    return "".join(str(label).split()).replace("　", "")


NAME_TO_CODE = {_norm(k): v for k, v in NAME_TO_CODE.items()}
NAME_TO_CODE.update({
    _norm("エネルギー kJ"): "ENERC",
    _norm("エネルギー kcal"): "ENERC_KCAL",
    _norm("βクリプトキサンチン"): "CRYPXB",
    "食物繊維総量": "FIB-",
    "レチノール活性当量": "VITA_RAE",
    "利用可能炭水化物（単糖当量）": "CHOAVLM",
    "差引き法による利用可能炭水化物": "CHOAVLDF-",
})


def item_by_code(conn, food_code):
    row = conn.execute(
        "SELECT id FROM items WHERE source='MEXT Standard Tables' AND source_url = ?",
        (f"mext_{food_code}",)).fetchone()
    return row[0] if row else None


def corrections_from_chapter2(wb):
    """(food_code, code, label, right) from the per-value sheet."""
    ws = wb["本表第2章"]
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[2]:
            continue
        food_code, item, right = str(row[2]).strip(), row[5], row[7]
        if not item or str(item).strip() in SKIP_ITEMS:
            continue
        label = str(item).strip()
        code = NAME_TO_CODE.get(_norm(label))
        if code:
            yield food_code, code, label, right


def corrections_from_full_rows(wb):
    """The 本表 sheet republishes entire rows. Columns are located by reading
    the 成分識別子 row rather than assuming the main table's offsets, because
    this sheet is shifted one column right."""
    ws = wb["本表"]
    rows = list(ws.iter_rows(values_only=True))
    code_row = next((i for i, r in enumerate(rows)
                     if any(str(c).strip() == "成分識別子" for c in r if c)), None)
    if code_row is None:
        return
    header = rows[code_row]
    col_of = {str(c).strip(): i for i, c in enumerate(header) if c}
    # 食品番号 sits in the block above; find it by the label row
    num_col = next((i for i, c in enumerate(rows[1]) if c and "食品番号" in str(c).replace("　", "")), 2)
    for r in rows[code_row + 1:]:
        if not r or not r[num_col]:
            continue
        # Each food appears twice, as 誤 (what was published) and 正 (the
        # correction). Only the correction is applied — taking every row and
        # letting the last one win would write the withdrawn values back, and
        # it worked here only because 正 happens to come second.
        if str(r[0]).strip() != "正":
            continue
        food_code = str(r[num_col]).strip()
        for _col, code, jp, _unit in NUTRIENT_COLS:
            # Located by 成分識別子, never by name: エネルギー appears twice
            # (kJ and kcal), and matching on the label wrote the kilojoule
            # figure into the kilocalorie field.
            i = col_of.get(code)
            if i is None or i >= len(r):
                continue
            yield food_code, code, jp, r[i]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    if not XLSX.is_file():
        sys.exit(f"{XLSX} not found — download MEXT's 正誤表 first")

    wb = openpyxl.load_workbook(XLSX, read_only=True)
    conn = sqlite3.connect(DB_PATH, timeout=60)
    conn.row_factory = sqlite3.Row

    applied = skipped = unchanged = 0
    for source_sheet, gen in (("本表第2章", corrections_from_chapter2(wb)),
                              ("本表", corrections_from_full_rows(wb))):
        for food_code, code, label, right in gen:
            item_id = item_by_code(conn, food_code)
            if not item_id:
                skipped += 1
                continue
            amount, quality = clean_value(right)
            if amount is None:
                skipped += 1
                continue
            cur = conn.execute(
                "SELECT id, amount, quality FROM nutrients WHERE item_id = ? AND code = ?",
                (item_id, code)).fetchone()
            if not cur:
                skipped += 1
                continue
            if abs((cur["amount"] or 0) - amount) < 1e-9 and cur["quality"] == quality:
                unchanged += 1
                continue
            print(f"  {food_code} {label} [{code}]: "
                  f"{cur['amount']} ({cur['quality']}) -> {amount} ({quality})   [{source_sheet}]")
            applied += 1
            if not args.dry_run:
                conn.execute("UPDATE nutrients SET amount = ?, quality = ? WHERE id = ?",
                             (amount, quality, cur["id"]))

    # The four headline figures are denormalised into `nutrition`; a corrected
    # value there has to follow, or the page and its own table disagree.
    macros = {"ENERC_KCAL": "energy_kcal", "PROT-": "protein_g",
              "FAT-": "fat_g", "CHOCDF-": "carbohydrate_g"}
    if not args.dry_run and applied:
        for code, field in macros.items():
            conn.execute(f"""
                UPDATE nutrition SET {field} = (
                    SELECT n.amount FROM nutrients n
                    WHERE n.item_id = nutrition.item_id AND n.code = ?)
                WHERE EXISTS (SELECT 1 FROM nutrients n
                              WHERE n.item_id = nutrition.item_id AND n.code = ?)""",
                         (code, code))
        conn.commit()
    conn.close()
    verb = "would apply" if args.dry_run else "applied"
    print(f"\n{verb}: {applied} corrections, {unchanged} already correct, {skipped} not applicable")


if __name__ == "__main__":
    main()
